from openpyxl import *
import os
import webbrowser
from datetime import datetime

def findlinkbykitab (kitab):
	kitabdbpath="~/Downloads/Kitab_Database.xlsx"
	kitabdbpath= os.path.expanduser(kitabdbpath) 
	wb1=load_workbook(kitabdbpath)
	sheet_obj1 = wb1.active

	if kitab == "Aalim Ghulam":
		cell_obj1 = sheet_obj1.cell(row = 1, column = 2)
	elif kitab == "Academic Writing":
		cell_obj1 = sheet_obj1.cell(row = 2, column = 2)
	elif kitab == "Adab Arabi":
		cell_obj1 = sheet_obj1.cell(row = 3, column = 2)
	elif kitab == "Adab Fatemi":
		cell_obj1 = sheet_obj1.cell(row = 4, column = 2)
	elif kitab == "Barnamaj":
		cell_obj1 = sheet_obj1.cell(row = 5, column = 2)
	elif kitab == "Emotional":
		cell_obj1 = sheet_obj1.cell(row = 6, column = 2)
	elif kitab == "Free Period":
		cell_obj1 = sheet_obj1.cell(row = 7, column = 2)
	elif kitab == "HCIW":
		cell_obj1 = sheet_obj1.cell(row = 8, column = 2)
	elif kitab == "Ikhwan":
		cell_obj1 = sheet_obj1.cell(row = 9, column = 2)
	elif kitab == "Language":
		cell_obj1 = sheet_obj1.cell(row = 10, column = 2)
	elif kitab == "Literature":
		cell_obj1 = sheet_obj1.cell(row = 11, column = 2)
	elif kitab == "Majalis":
		cell_obj1 = sheet_obj1.cell(row = 12, column = 2)
	elif kitab == "Management":
		cell_obj1 = sheet_obj1.cell(row = 13, column = 2)
	elif kitab == "Maqamat":
		cell_obj1 = sheet_obj1.cell(row = 14, column = 2)
	elif kitab == "Maqraat":
		cell_obj1 = sheet_obj1.cell(row = 15, column = 2)
	elif kitab == "Masool":
		cell_obj1 = sheet_obj1.cell(row = 16, column = 2)
	elif kitab == "Mukhtasar":
		cell_obj1 = sheet_obj1.cell(row = 17, column = 2)
	elif kitab == "Muntakhaba":
		cell_obj1 = sheet_obj1.cell(row = 18, column = 2)
	elif kitab == "Nehj":
		cell_obj1 = sheet_obj1.cell(row = 19, column = 2)
	elif kitab == "Risala Alif":
		cell_obj1 = sheet_obj1.cell(row = 20, column = 2)
	elif kitab == "Risala B":
		cell_obj1 = sheet_obj1.cell(row = 21, column = 2)
	elif kitab == "Takhassus":
		cell_obj1 = sheet_obj1.cell(row = 22, column = 2)
	elif kitab == "Uloom Quran":
		cell_obj1 = sheet_obj1.cell(row = 23, column = 2)
	elif kitab == "Uyun":
		cell_obj1 = sheet_obj1.cell(row = 24, column = 2)
	else:
		print ("Something went wrong")
	link=(cell_obj1.value)

	webbrowser.open_new (link)


def findpreviouspages (kitab):
	

	nishanipath="~/Downloads/Nishani.xlsx"
	nishanipath=os.path.expanduser(nishanipath)
	wb2=load_workbook(nishanipath)
	sheet_obj2 = wb2.active
	if kitab == "Aalim Ghulam":
		cell_obj2 = sheet_obj2.cell(row = 1, column = 2)
		notes=sheet_obj2.cell(row=1, column=3)
	elif kitab == "Academic Writing":
		cell_obj2 = sheet_obj2.cell(row = 2, column = 2)
		notes=sheet_obj2.cell(row=2, column=3)
	elif kitab == "Adab Arabi":
		cell_obj2 = sheet_obj2.cell(row = 3, column = 2)
		notes=sheet_obj2.cell(row=3, column=3)
	elif kitab == "Adab Fatemi":
		cell_obj2 = sheet_obj2.cell(row = 4, column = 2)
		notes=sheet_obj2.cell(row=4, column=3)
	elif kitab == "Barnamaj":
		cell_obj2 = sheet_obj2.cell(row = 5, column = 2)
		notes=sheet_obj2.cell(row=5, column=3)
	elif kitab == "Emotional":
		cell_obj2 = sheet_obj2.cell(row = 6, column = 2)
		notes=sheet_obj2.cell(row=6, column=3)
	elif kitab == "Free Period":
		cell_obj2 = sheet_obj2.cell(row = 7, column = 2)
		notes=sheet_obj2.cell(row=7, column=3)
	elif kitab == "HCIW":
		cell_obj2 = sheet_obj2.cell(row = 8, column = 2)
		notes=sheet_obj2.cell(row=8, column=3)
	elif kitab == "Ikhwan":
		cell_obj2 = sheet_obj2.cell(row = 9, column = 2)
		notes=sheet_obj2.cell(row=9, column=3)
	elif kitab == "Language":
		cell_obj2 = sheet_obj2.cell(row = 10, column = 2)
		notes=sheet_obj2.cell(row=10, column=3)
	elif kitab == "Literature":
		cell_obj2 = sheet_obj2.cell(row = 11, column = 2)
		notes=sheet_obj2.cell(row=11, column=3)
	elif kitab == "Majalis":
		cell_obj2 = sheet_obj2.cell(row = 12, column = 2)
		notes=sheet_obj2.cell(row=12, column=3)
	elif kitab == "Management":
		cell_obj2 = sheet_obj2.cell(row = 13, column = 2)
		notes=sheet_obj2.cell(row=13, column=3)
	elif kitab == "Maqamat":
		cell_obj2 = sheet_obj2.cell(row = 14, column = 2)
		notes=sheet_obj2.cell(row=14, column=3)
	elif kitab == "Maqraat":
		cell_obj2 = sheet_obj2.cell(row = 15, column = 2)
		notes=sheet_obj2.cell(row=15, column=3)
	elif kitab == "Masool":
		cell_obj2 = sheet_obj2.cell(row = 16, column = 2)
		notes=sheet_obj2.cell(row=16, column=3)
	elif kitab == "Mukhtasar":
		cell_obj2 = sheet_obj2.cell(row = 17, column = 2)
		notes=sheet_obj2.cell(row=17, column=3)
	elif kitab == "Muntakhaba":
		cell_obj2 = sheet_obj2.cell(row = 18, column = 2)
		notes=sheet_obj2.cell(row=18, column=3)
	elif kitab == "Nehj":
		cell_obj2 = sheet_obj2.cell(row = 19, column = 2)
		notes=sheet_obj2.cell(row=19, column=3)
	elif kitab == "Risala Alif":
		cell_obj2 = sheet_obj2.cell(row = 20, column = 2)
		notes=sheet_obj2.cell(row=20, column=3)
	elif kitab == "Risala B":
		cell_obj2 = sheet_obj2.cell(row = 21, column = 2)
		notes=sheet_obj2.cell(row=21, column=3)
	elif kitab == "Takhassus":
		cell_obj2 = sheet_obj2.cell(row = 22, column = 2)
		notes=sheet_obj2.cell(row=22, column=3)
	elif kitab == "Uloom Quran":
		cell_obj2 = sheet_obj2.cell(row = 23, column = 2)
		notes=sheet_obj2.cell(row=23, column=3)
	elif kitab == "Uyun":
		cell_obj2 = sheet_obj2.cell(row = 24, column = 2)
		notes=sheet_obj2.cell(row=24, column=3)
	else:
		print ("Something went wrong")

	prevpage=(cell_obj2.value)
	notes = (notes.value)
	print ("In {}, you last reached {}".format(kitab, prevpage))
	c= "Your notes for {}, were: {}".format(kitab, notes)
	print (c)


def savequestions (kitab,questions):
	listofkitabs={"Aalim Ghulam":"D1", "Academic Writing":"D2", "Adab Arabi":"D3", "Adab Fatemi":"D4", "Barnamaj": "D5", "Emotional":"D6", "Free Period":"D7", 
	"HCIW":"D8", "Ikhwan": "D9", "Language":"D10", "Literature":"D11", "Majalis":"D12", "Management":"D13", "Maqamat":"D14","Maqraat":"D15", "Masool":"D16","Mukhtasar":"D17", 
	"Muntakhaba":"D18", "Nehj":"D19", "Risala Alif": "D20", "Risala B":"D21", "Takhassus":"D22", "Uloom Quran":"D23","Uyun":"D24"
	}
	from openpyxl import load_workbook
	nishanipath="~/Downloads/Nishani.xlsx"
	nishanipath=os.path.expanduser(nishanipath)
	workbook = load_workbook(nishanipath)
	sheet_nishani = workbook.active
	sheet_nishani [listofkitabs.get(subject)] = questions
	workbook.save(nishanipath)

schedpath = "~/Downloads/Schedule.xlsx"
schedpath=os.path.expanduser(schedpath)
wb_obj = openpyxl.load_workbook(schedpath)
sheet_obj = wb_obj.active
daynum=datetime.today().weekday()
daynum=daynum+2
periodindex = [2,3,4,5,6,7,8,9,10]
for index in periodindex:
	cell_obj = sheet_obj.cell(row = index, column = daynum)
	subject = cell_obj.value
	findpreviouspages(subject)
	findlinkbykitab (subject)
	questions=input("Enter any questions you have about this period. Otherwise, press Enter. ")
	savequestions (kitab,questions)


