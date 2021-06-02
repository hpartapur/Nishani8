from openpyxl import *
import os
def findlinkbykitab (kitab):
	kitabdbpath="~/Desktop/Nishani8/Kitab_Database.xlsx"
	kitabdbpath= os.path.expanduser(kitabdbpath) 
	wb1=load_workbook(kitabdbpath)
	sheet_obj1 = wb1.active

	if kitab == "Aalim Ghulam":
		cell_obj1 = sheet_obj1.cell(row = 1, column = 2)
	elif kitab == "Academic Writing":
		cell_obj1 = sheet_obj1.cell(row = 2, column = 2)
	elif kitab == "Adab Arabi":
		cell_obj1 = sheet_obj1.cell(row = 3, column = 2)
	elif kitab == "Adab Fatemi":
		cell_obj1 = sheet_obj1.cell(row = 4, column = 2)
	elif kitab == "Barnamaj":
		cell_obj1 = sheet_obj1.cell(row = 5, column = 2)
	elif kitab == "Emotional":
		cell_obj1 = sheet_obj1.cell(row = 6, column = 2)
	elif kitab == "Free Period":
		cell_obj1 = sheet_obj1.cell(row = 7, column = 2)
	elif kitab == "HCIW":
		cell_obj1 = sheet_obj1.cell(row = 8, column = 2)
	elif kitab == "Ikhwan":
		cell_obj1 = sheet_obj1.cell(row = 9, column = 2)
	elif kitab == "Language":
		cell_obj1 = sheet_obj1.cell(row = 10, column = 2)
	elif kitab == "Literature":
		cell_obj1 = sheet_obj1.cell(row = 11, column = 2)
	elif kitab == "Majalis":
		cell_obj1 = sheet_obj1.cell(row = 12, column = 2)
	elif kitab == "Management":
		cell_obj1 = sheet_obj1.cell(row = 13, column = 2)
	elif kitab == "Maqamat":
		cell_obj1 = sheet_obj1.cell(row = 14, column = 2)
	elif kitab == "Maqrat":
		cell_obj1 = sheet_obj1.cell(row = 15, column = 2)
	elif kitab == "Masool":
		cell_obj1 = sheet_obj1.cell(row = 16, column = 2)
	elif kitab == "Mukhtasar":
		cell_obj1 = sheet_obj1.cell(row = 17, column = 2)
	elif kitab == "Muntakhaba":
		cell_obj1 = sheet_obj1.cell(row = 18, column = 2)
	elif kitab == "Nehj":
		cell_obj1 = sheet_obj1.cell(row = 19, column = 2)
	elif kitab == "Risala Alif":
		cell_obj1 = sheet_obj1.cell(row = 20, column = 2)
	elif kitab == "Risala B":
		cell_obj1 = sheet_obj1.cell(row = 21, column = 2)
	elif kitab == "Takhassus":
		cell_obj1 = sheet_obj1.cell(row = 22, column = 2)
	elif kitab == "Uloom Quran":
		cell_obj1 = sheet_obj1.cell(row = 23, column = 2)
	elif kitab == "Uyun":
		cell_obj1 = sheet_obj1.cell(row = 24, column = 2)
	else:
		print ("Something went wrong")
	link=(cell_obj1.value)

	webbrowser.open_new (link)

def savepagenumber (subject, newnumber,notes):
	listofkitabs={"Aalim Ghulam":"B1", "Academic Writing":"B2", "Adab Arabi":"B3", "Adab Fatemi":"B4", "Barnamaj": "B5", "Emotional":"B6", "Free Period":"B7", 
	"HCIW":"B8", "Ikhwan": "B9", "Language":"B10", "Literature":"B11", "Majalis":"B12", "Management":"B13", "Maqamat":"B14","Maqrat":"B15", "Masool":"B16","Mukhtasar":"B17", 
	"Muntakhaba":"B18", "Nehj":"B19", "Risala Alif": "B20", "Risala B":"B21", "Takhassus":"B22", "Uloom Quran":"B23","Uyun":"B24"
	}
	listofkitabs2={"Aalim Ghulam":"C1", "Academic Writing":"C2", "Adab Arabi":"C3", "Adab Fatemi":"C4", "Barnamaj": "C5", "Emotional":"C6", "Free Period":"C7", 
	"HCIW":"C8", "Ikhwan": "C9", "Language":"C10", "Literature":"C11", "Majalis":"C12", "Management":"C13", "Maqamat":"C14","Maqrat":"C15", "Masool":"C16","Mukhtasar":"C17", 
	"Muntakhaba":"C18", "Nehj":"C19", "Risala Alif": "C20", "Risala B":"C21", "Takhassus":"C22", "Uloom Quran":"C23","Uyun":"C24"
	}
	from openpyxl import load_workbook
	nishanipath="~/Desktop/Nishani8//Nishani.xlsx"
	nishanipath=os.path.expanduser(nishanipath)
	workbook = load_workbook(nishanipath)
	sheet_nishani = workbook.active
	sheet_nishani [listofkitabs.get(subject)] = newnumber
	sheet_nishani [listofkitabs2.get(subject)] = notes
	workbook.save(nishanipath)

def findpreviouspages (kitab):
	

	nishanipath="~/Desktop/Nishani8/Nishani.xlsx"
	nishanipath=os.path.expanduser(nishanipath)
	wb2=load_workbook(nishanipath)
	sheet_obj2 = wb2.active
	if kitab == "Aalim Ghulam":
		cell_obj2 = sheet_obj2.cell(row = 1, column = 2)
		notes=sheet_obj2.cell(row=1, column=3)
		questions=sheet_obj2.cell(row=1, column=4)
	elif kitab == "Academic Writing":
		cell_obj2 = sheet_obj2.cell(row = 2, column = 2)
		notes=sheet_obj2.cell(row=2, column=3)
		questions=sheet_obj2.cell(row=2, column=4)
	elif kitab == "Adab Arabi":
		cell_obj2 = sheet_obj2.cell(row = 3, column = 2)
		notes=sheet_obj2.cell(row=3, column=3)
		questions=sheet_obj2.cell(row=3, column=4)
	elif kitab == "Adab Fatemi":
		cell_obj2 = sheet_obj2.cell(row = 4, column = 2)
		notes=sheet_obj2.cell(row=4, column=3)
		questions=sheet_obj2.cell(row=4, column=4)
	elif kitab == "Barnamaj":
		cell_obj2 = sheet_obj2.cell(row = 5, column = 2)
		notes=sheet_obj2.cell(row=5, column=3)
		questions=sheet_obj2.cell(row=5, column=4)
	elif kitab == "Emotional":
		cell_obj2 = sheet_obj2.cell(row = 6, column = 2)
		notes=sheet_obj2.cell(row=6, column=3)
		questions=sheet_obj2.cell(row=6, column=4)
	elif kitab == "Free Period":
		cell_obj2 = sheet_obj2.cell(row = 7, column = 2)
		notes=sheet_obj2.cell(row=7, column=3)
		questions=sheet_obj2.cell(row=7, column=4)
	elif kitab == "HCIW":
		cell_obj2 = sheet_obj2.cell(row = 8, column = 2)
		notes=sheet_obj2.cell(row=8, column=3)
		questions=sheet_obj2.cell(row=8, column=4)
	elif kitab == "Ikhwan":
		cell_obj2 = sheet_obj2.cell(row = 9, column = 2)
		notes=sheet_obj2.cell(row=9, column=3)
		questions=sheet_obj2.cell(row=9, column=4)
	elif kitab == "Language":
		cell_obj2 = sheet_obj2.cell(row = 10, column = 2)
		notes=sheet_obj2.cell(row=10, column=3)
		questions=sheet_obj2.cell(row=10, column=4)
	elif kitab == "Literature":
		cell_obj2 = sheet_obj2.cell(row = 11, column = 2)
		notes=sheet_obj2.cell(row=11, column=3)
		questions=sheet_obj2.cell(row=11, column=4)
	elif kitab == "Majalis":
		cell_obj2 = sheet_obj2.cell(row = 12, column = 2)
		notes=sheet_obj2.cell(row=12, column=3)
		questions=sheet_obj2.cell(row=12, column=4)
	elif kitab == "Management":
		cell_obj2 = sheet_obj2.cell(row = 13, column = 2)
		notes=sheet_obj2.cell(row=13, column=3)
		questions=sheet_obj2.cell(row=13, column=4)
	elif kitab == "Maqamat":
		cell_obj2 = sheet_obj2.cell(row = 14, column = 2)
		notes=sheet_obj2.cell(row=14, column=3)
		questions=sheet_obj2.cell(row=14, column=4)
	elif kitab == "Maqrat":
		cell_obj2 = sheet_obj2.cell(row = 15, column = 2)
		notes=sheet_obj2.cell(row=15, column=3)
		questions=sheet_obj2.cell(row=15, column=4)
	elif kitab == "Masool":
		cell_obj2 = sheet_obj2.cell(row = 16, column = 2)
		notes=sheet_obj2.cell(row=16, column=3)
		questions=sheet_obj2.cell(row=16, column=4)
	elif kitab == "Mukhtasar":
		cell_obj2 = sheet_obj2.cell(row = 17, column = 2)
		notes=sheet_obj2.cell(row=17, column=3)
		questions=sheet_obj2.cell(row=17, column=4)
	elif kitab == "Muntakhaba":
		cell_obj2 = sheet_obj2.cell(row = 18, column = 2)
		notes=sheet_obj2.cell(row=18, column=3)
		questions=sheet_obj2.cell(row=18, column=4)
	elif kitab == "Nehj":
		cell_obj2 = sheet_obj2.cell(row = 19, column = 2)
		notes=sheet_obj2.cell(row=19, column=3)
		questions=sheet_obj2.cell(row=19, column=4)
	elif kitab == "Risala Alif":
		cell_obj2 = sheet_obj2.cell(row = 20, column = 2)
		notes=sheet_obj2.cell(row=20, column=3)
		questions=sheet_obj2.cell(row=20, column=4)
	elif kitab == "Risala B":
		cell_obj2 = sheet_obj2.cell(row = 21, column = 2)
		notes=sheet_obj2.cell(row=21, column=3)
		questions=sheet_obj2.cell(row=21, column=4)
	elif kitab == "Takhassus":
		cell_obj2 = sheet_obj2.cell(row = 22, column = 2)
		notes=sheet_obj2.cell(row=22, column=3)
		questions=sheet_obj2.cell(row=22, column=4)
	elif kitab == "Uloom Quran":
		cell_obj2 = sheet_obj2.cell(row = 23, column = 2)
		notes=sheet_obj2.cell(row=23, column=3)
		questions=sheet_obj2.cell(row=23, column=4)
	elif kitab == "Uyun":
		cell_obj2 = sheet_obj2.cell(row = 24, column = 2)
		notes=sheet_obj2.cell(row=24, column=3)
		questions=sheet_obj2.cell(row=24, column=4)
	else:
		print ("Something went wrong")


	prevpage=(cell_obj2.value)
	notes = (notes.value)
	questions=(questions.value)
	print ()
	print ()
	print ()
	print ()
	print (kitab)
	print ()
	print ("In {}, you last reached {}".format(kitab, prevpage))
	print()
	c= "Your notes for {}, were: {}".format(kitab, notes)
	print (c)
	print ()
	print ("Your questions from iadat were ")
	print (questions)

import webbrowser
webbrowser.open ('https://www.jameasaifiyah.org/Students/stuLogin.aspx')
from datetime import datetime
import openpyxl
schedpath = "~/Desktop/Nishani8/Schedule.xlsx"
schedpath=os.path.expanduser(schedpath)
wb_obj = openpyxl.load_workbook(schedpath)
sheet_obj = wb_obj.active
daynum=datetime.today().weekday()
daynum=daynum+2


periodindex = [2,3,4,5,6,7,8,9,10]
for index in periodindex:
	cell_obj = sheet_obj.cell(row = index, column = daynum)
	subject = cell_obj.value
	findpreviouspages(subject)
	findlinkbykitab (subject)
	notes=input("Enter any notes or remarks you want to save for this period. Otherwise, just press Enter. ")
	print ()
	pagenumber=input ("Enter the page number you reached in this period. ")
	savepagenumber(subject, pagenumber, notes)
	




